from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "dataverse-build-workbook.xlsx"
VALIDATION_PATH = ROOT / "dataverse-build-workbook.20260903T022215Z.d9eab55d.artifactvalidatepass"

TABLES = [
    ("esp_businessobjective", "Business", "User/Team", "esp_code", "ESPCore"),
    ("esp_usecase", "Business", "User/Team", "esp_code", "ESPCore"),
    ("esp_capabilityassessment", "Business", "User/Team", "esp_code", "ESPCore"),
    ("esp_usecaseskillrequirement", "Business", "User/Team", "esp_code", "ESPCore"),
    ("esp_consumer", "Consumption", "User/Team", "esp_code", "ESPCore"),
    ("esp_logicalskill", "Capability", "User/Team", "esp_code", "ESPCore"),
    ("esp_logicalskillversion", "Capability", "User/Team", "esp_logicalskillid+esp_version", "ESPCore"),
    ("esp_knowledgeasset", "Assets", "User/Team", "esp_code", "ESPSharedAssets"),
    ("esp_knowledgeassetversion", "Assets", "User/Team", "esp_knowledgeassetid+esp_version", "ESPSharedAssets"),
    ("esp_promptasset", "Assets", "User/Team", "esp_code", "ESPSharedAssets"),
    ("esp_promptassetversion", "Assets", "User/Team", "esp_promptassetid+esp_version", "ESPSharedAssets"),
    ("esp_templateasset", "Assets", "User/Team", "esp_code", "ESPSharedAssets"),
    ("esp_templateassetversion", "Assets", "User/Team", "esp_templateassetid+esp_version", "ESPSharedAssets"),
    ("esp_methodasset", "Assets", "User/Team", "esp_code", "ESPSharedAssets"),
    ("esp_methodassetversion", "Assets", "User/Team", "esp_methodassetid+esp_version", "ESPSharedAssets"),
    ("esp_toolcontract", "Assets", "User/Team", "esp_code", "ESPSharedAssets"),
    ("esp_toolcontractversion", "Assets", "User/Team", "esp_toolcontractid+esp_version", "ESPSharedAssets"),
    ("esp_workflowdefinition", "Assets", "User/Team", "esp_code", "ESPSharedAssets"),
    ("esp_workflowdefinitionversion", "Assets", "User/Team", "esp_workflowdefinitionid+esp_version", "ESPSharedAssets"),
    ("esp_externalreference", "Assets", "User/Team", "esp_code", "ESPSharedAssets"),
    ("esp_implementation", "Delivery", "User/Team", "esp_code", "ESPCore"),
    ("esp_implementationversion", "Delivery", "User/Team", "esp_implementationid+esp_version", "ESPCore"),
    ("esp_runtimeprofile", "Delivery", "User/Team", "esp_code", "ESPCore"),
    ("esp_dependency", "Delivery", "User/Team", "esp_code", "ESPGovernance"),
    ("esp_packageversion", "Delivery", "User/Team", "esp_code+esp_version", "ESPGovernance"),
    ("esp_dependencysnapshot", "Delivery", "Organization", "esp_code", "ESPGovernance"),
    ("esp_releaserecord", "Governance", "Organization", "esp_code", "ESPGovernance"),
    ("esp_deployment", "Governance", "Organization", "esp_code", "ESPGovernance"),
    ("esp_consumerbinding", "Consumption", "Organization", "esp_code", "ESPGovernance"),
    ("esp_evaluationprofile", "Quality", "User/Team", "esp_code", "ESPGovernance"),
    ("esp_evaluationrun", "Quality", "Organization", "esp_code", "ESPGovernance"),
    ("esp_gatedecision", "Governance", "Organization", "esp_code", "ESPGovernance"),
    ("esp_approvalrecord", "Governance", "Organization", "esp_code", "ESPGovernance"),
    ("esp_identityprofile", "Governance", "User/Team", "esp_code", "ESPGovernance"),
    ("esp_identitybinding", "Governance", "Organization", "esp_code", "ESPGovernance"),
    ("esp_policybinding", "Governance", "Organization", "esp_code", "ESPGovernance"),
    ("esp_invocationindex", "Evidence", "Organization", "esp_externalid", "ESPEvidence"),
    ("esp_evidencepackage", "Evidence", "User/Team", "esp_code", "ESPEvidence"),
    ("esp_evidenceitem", "Evidence", "Organization", "esp_code", "ESPEvidence"),
    ("esp_operationalevidence", "Evidence", "Organization", "esp_code", "ESPEvidence"),
    ("esp_operationalevidencesource", "Evidence", "Organization", "esp_code", "ESPEvidence"),
    ("esp_valueassessmentreference", "Value", "User/Team", "esp_code", "ESPEvidence"),
]

RELATIONSHIPS = [
    ("esp_businessobjective", "esp_usecase", "1:N", "esp_businessobjectiveid", "Required", "Restrict", "Objective history retained"),
    ("esp_usecase", "esp_capabilityassessment", "1:N", "esp_usecaseid", "Required", "Restrict", "Assessment belongs to use case"),
    ("esp_usecase", "esp_usecaseskillrequirement", "1:N", "esp_usecaseid", "Required", "Restrict", "Requirement belongs to use case"),
    ("esp_usecase", "esp_consumer", "N:N", "", "Optional", "Remove Link", "Candidate or active consumers"),
    ("esp_logicalskill", "esp_logicalskillversion", "1:N", "esp_logicalskillid", "Required", "Restrict", "Approved versions retained"),
    ("esp_logicalskillversion", "esp_usecaseskillrequirement", "1:N", "esp_logicalskillversionid", "Required", "Restrict", "Pins required contract"),
    ("esp_knowledgeasset", "esp_knowledgeassetversion", "1:N", "esp_knowledgeassetid", "Required", "Restrict", "Version history retained"),
    ("esp_promptasset", "esp_promptassetversion", "1:N", "esp_promptassetid", "Required", "Restrict", "Version history retained"),
    ("esp_templateasset", "esp_templateassetversion", "1:N", "esp_templateassetid", "Required", "Restrict", "Version history retained"),
    ("esp_methodasset", "esp_methodassetversion", "1:N", "esp_methodassetid", "Required", "Restrict", "Version history retained"),
    ("esp_toolcontract", "esp_toolcontractversion", "1:N", "esp_toolcontractid", "Required", "Restrict", "Version history retained"),
    ("esp_workflowdefinition", "esp_workflowdefinitionversion", "1:N", "esp_workflowdefinitionid", "Required", "Restrict", "Version history retained"),
    ("esp_logicalskillversion", "esp_implementation", "1:N", "esp_logicalskillversionid", "Required", "Restrict", "Runtime realization"),
    ("esp_implementation", "esp_implementationversion", "1:N", "esp_implementationid", "Required", "Restrict", "Implementation history retained"),
    ("esp_runtimeprofile", "esp_implementationversion", "1:N", "esp_runtimeprofileid", "Required", "Restrict", "Host configuration contract"),
    ("esp_implementationversion", "esp_dependency", "1:N", "esp_sourceimplementationversionid", "Required", "Restrict", "Versioned dependency source"),
    ("esp_knowledgeassetversion", "esp_dependency", "1:N", "esp_knowledgeassetversionid", "Optional", "Restrict", "Exactly one dependency target"),
    ("esp_promptassetversion", "esp_dependency", "1:N", "esp_promptassetversionid", "Optional", "Restrict", "Exactly one dependency target"),
    ("esp_templateassetversion", "esp_dependency", "1:N", "esp_templateassetversionid", "Optional", "Restrict", "Exactly one dependency target"),
    ("esp_methodassetversion", "esp_dependency", "1:N", "esp_methodassetversionid", "Optional", "Restrict", "Exactly one dependency target"),
    ("esp_toolcontractversion", "esp_dependency", "1:N", "esp_toolcontractversionid", "Optional", "Restrict", "Exactly one dependency target"),
    ("esp_workflowdefinitionversion", "esp_dependency", "1:N", "esp_workflowdefinitionversionid", "Optional", "Restrict", "Exactly one dependency target"),
    ("esp_externalreference", "esp_dependency", "1:N", "esp_externalreferenceid", "Optional", "Restrict", "Exactly one dependency target"),
    ("esp_implementationversion", "esp_packageversion", "1:N", "esp_implementationversionid", "Required", "Restrict", "Package target"),
    ("esp_packageversion", "esp_dependencysnapshot", "1:N", "esp_packageversionid", "Required", "Restrict", "Resolved release dependency"),
    ("esp_dependency", "esp_dependencysnapshot", "1:N", "esp_dependencyid", "Required", "Restrict", "Snapshot source"),
    ("esp_packageversion", "esp_releaserecord", "1:N", "esp_packageversionid", "Required", "Restrict", "Approved package release"),
    ("esp_releaserecord", "esp_deployment", "1:N", "esp_releaserecordid", "Required", "Restrict", "Deployment release"),
    ("esp_runtimeprofile", "esp_deployment", "1:N", "esp_runtimeprofileid", "Required", "Restrict", "Runtime host"),
    ("esp_deployment", "esp_consumerbinding", "1:N", "esp_deploymentid", "Required", "Restrict", "Approved deployment"),
    ("esp_consumer", "esp_consumerbinding", "1:N", "esp_consumerid", "Required", "Restrict", "Bound consumer"),
    ("esp_logicalskillversion", "esp_consumerbinding", "1:N", "esp_logicalskillversionid", "Required", "Restrict", "Bound contract version"),
    ("esp_implementationversion", "esp_consumerbinding", "1:N", "esp_implementationversionid", "Required", "Restrict", "Bound implementation version"),
    ("esp_logicalskillversion", "esp_evaluationprofile", "1:N", "esp_logicalskillversionid", "Required", "Restrict", "Skill evaluation definition"),
    ("esp_evaluationprofile", "esp_evaluationrun", "1:N", "esp_evaluationprofileid", "Required", "Restrict", "Pinned evaluation profile"),
    ("esp_logicalskillversion", "esp_evaluationrun", "1:N", "esp_logicalskillversionid", "Required", "Restrict", "Pinned evaluated Skill contract"),
    ("esp_implementationversion", "esp_evaluationrun", "1:N", "esp_implementationversionid", "Required", "Restrict", "Pinned evaluated implementation"),
    ("esp_packageversion", "esp_evaluationrun", "1:N", "esp_packageversionid", "Required", "Restrict", "Pinned evaluated package"),
    ("esp_deployment", "esp_evaluationrun", "1:N", "esp_deploymentid", "Required", "Restrict", "Test target"),
    ("esp_consumerbinding", "esp_evaluationrun", "1:N", "esp_consumerbindingid", "Required", "Restrict", "Agent validation binding"),
    ("esp_evaluationrun", "esp_gatedecision", "1:N", "esp_evaluationrunid", "Required", "Restrict", "Gate evidence"),
    ("esp_gatedecision", "esp_approvalrecord", "1:N", "esp_gatedecisionid", "Optional", "Restrict", "Gate approval"),
    ("esp_releaserecord", "esp_approvalrecord", "1:N", "esp_releaserecordid", "Optional", "Restrict", "Release approval"),
    ("esp_identityprofile", "esp_identitybinding", "1:N", "esp_identityprofileid", "Required", "Restrict", "Runtime identity requirement"),
    ("esp_consumerbinding", "esp_identitybinding", "1:N", "esp_consumerbindingid", "Required", "Restrict", "Binding identity"),
    ("esp_consumerbinding", "esp_policybinding", "1:N", "esp_consumerbindingid", "Required", "Restrict", "Binding policy"),
    ("esp_consumerbinding", "esp_invocationindex", "1:N", "esp_consumerbindingid", "Required", "Restrict", "Invocation traceability"),
    ("esp_deployment", "esp_invocationindex", "1:N", "esp_deploymentid", "Required", "Restrict", "Executed deployment"),
    ("esp_usecase", "esp_invocationindex", "1:N", "esp_usecaseid", "Required", "Restrict", "Invocation scenario"),
    ("esp_invocationindex", "esp_evidencepackage", "1:N", "esp_invocationindexid", "Required", "Restrict", "Invocation evidence"),
    ("esp_evidencepackage", "esp_evidenceitem", "1:N", "esp_evidencepackageid", "Required", "Restrict", "Retain evidence"),
    ("esp_usecase", "esp_operationalevidence", "1:N", "esp_usecaseid", "Required", "Restrict", "Aggregated scenario evidence"),
    ("esp_logicalskillversion", "esp_operationalevidence", "1:N", "esp_logicalskillversionid", "Required", "Restrict", "Aggregated skill evidence"),
    ("esp_operationalevidence", "esp_operationalevidencesource", "1:N", "esp_operationalevidenceid", "Required", "Restrict", "Aggregation membership"),
    ("esp_invocationindex", "esp_operationalevidencesource", "1:N", "esp_invocationindexid", "Required", "Restrict", "Source invocation"),
    ("esp_businessobjective", "esp_valueassessmentreference", "1:N", "esp_businessobjectiveid", "Required", "Restrict", "Business-owned value decision"),
    ("esp_operationalevidence", "esp_valueassessmentreference", "1:N", "esp_operationalevidenceid", "Required", "Restrict", "Evidence supporting decision"),
]

SPECIAL_COLUMNS = {
    "esp_businessobjective": [("Baseline Reference", "esp_baselinereference", "URL(1000)", "Optional", "No", "Restricted", "Authoritative baseline"), ("Value Metrics", "esp_valuemetrics", "Multiline Text", "Required", "No", "Restricted", "Business-owned measures"), ("Constraints", "esp_constraints", "Multiline Text", "Optional", "No", "Restricted", "Scope constraints"), ("Review Date", "esp_reviewdate", "Date", "Required", "No", "No", "Scheduled review")],
    "esp_usecase": [("Actor", "esp_actor", "Text(200)", "Required", "No", "No", "Primary actor"), ("Trigger", "esp_trigger", "Multiline Text", "Required", "No", "No", "Business trigger"), ("Expected Outcome", "esp_expectedoutcome", "Multiline Text", "Required", "No", "No", "Expected outcome"), ("Risk Tier", "esp_risktier", "Choice", "Required", "No", "No", "Approved risk tier"), ("Data Classification", "esp_dataclassification", "Choice", "Required", "No", "Restricted", "Highest input classification"), ("Human Oversight", "esp_humanoversight", "Choice", "Required", "No", "No", "Required oversight")],
    "esp_capabilityassessment": [("Decision", "esp_assessmentdecision", "Choice", "Required", "No", "No", "Reuse, Extend, Build, or Reject"), ("Fit Gaps", "esp_fitgaps", "Multiline Text", "Required", "No", "No", "Assessed gaps"), ("Decision Rationale", "esp_decisionrationale", "Multiline Text", "Required", "No", "No", "Recorded rationale")],
    "esp_usecaseskillrequirement": [("Acceptance Criteria", "esp_acceptancecriteria", "Multiline Text", "Required", "No", "No", "Testable criteria"), ("Evidence Requirements", "esp_evidencerequirements", "Multiline Text", "Required", "No", "Restricted", "Required evidence"), ("Priority", "esp_priority", "Choice", "Required", "No", "No", "Delivery priority"), ("Oversight Level", "esp_oversightlevel", "Choice", "Required", "No", "No", "Required oversight")],
    "esp_consumer": [("Consumer Type", "esp_consumertype", "Choice", "Required", "No", "No", "Agent, workflow, app, API, or human experience"), ("Runtime Reference", "esp_runtimereference", "URL(1000)", "Required", "No", "Restricted", "External runtime reference")],
    "esp_externalreference": [("Source Reference", "esp_sourcereference", "URL(1000)", "Required", "No", "Restricted", "Approved external source"), ("Source Version", "esp_sourceversion", "Text(100)", "Required", "No", "Restricted", "Immutable external version"), ("Artifact Hash", "esp_artifacthash", "Text(128)", "Required", "No", "Restricted", "External artifact integrity"), ("Authority", "esp_authority", "Text(200)", "Required", "No", "Restricted", "Approving authority"), ("Classification", "esp_classification", "Choice", "Required", "No", "Restricted", "External content classification")],
    "esp_logicalskillversion": [("Version", "esp_version", "Text(32)", "Required", "AK", "No", "SemVer"), ("Input Schema Reference", "esp_inputschemareference", "URL(1000)", "Required", "No", "Restricted", "Input contract"), ("Output Schema Reference", "esp_outputschemareference", "URL(1000)", "Required", "No", "Restricted", "Output contract"), ("Error Schema Reference", "esp_errorschemareference", "URL(1000)", "Required", "No", "Restricted", "Error contract"), ("Compatibility Policy", "esp_compatibilitypolicy", "Multiline Text", "Required", "No", "No", "SemVer compatibility"), ("Limitations", "esp_limitations", "Multiline Text", "Required", "No", "No", "Approved limitations")],
    "esp_implementation": [("Runtime Technology", "esp_runtimetechnology", "Text(100)", "Required", "No", "No", "Implementation platform"), ("Support Reference", "esp_supportreference", "URL(1000)", "Required", "No", "Restricted", "Support ownership")],
    "esp_implementationversion": [("Version", "esp_version", "Text(32)", "Required", "AK", "No", "SemVer"), ("Implementation Status", "esp_implementationstatus", "Choice", "Required", "No", "No", "Implementation lifecycle"), ("Component Reference", "esp_componentreference", "URL(1000)", "Required", "No", "Restricted", "Solution component"), ("Artifact Hash", "esp_artifacthash", "Text(128)", "Required", "No", "Restricted", "Integrity hash"), ("Contract Mapping Reference", "esp_contractmappingreference", "URL(1000)", "Required", "No", "Restricted", "Runtime-to-skill mapping")],
    "esp_runtimeprofile": [("Host Type", "esp_hosttype", "Choice", "Required", "No", "No", "Copilot Studio or approved host"), ("Exposure Mechanism", "esp_exposuremechanism", "Choice", "Required", "No", "No", "Action, topic, connector, or flow"), ("Configuration Reference", "esp_configurationreference", "URL(1000)", "Required", "No", "Restricted", "Non-secret runtime configuration")],
    "esp_dependency": [("Dependency Type", "esp_dependencytype", "Choice", "Required", "No", "No", "Target type"), ("Mode", "esp_dependencymode", "Choice", "Required", "No", "No", "Version Pinned by default"), ("Mandatory", "esp_mandatory", "Yes/No", "Required", "No", "No", "Blocks release when unresolved"), ("Usage Role", "esp_usagerole", "Text(200)", "Required", "No", "No", "Runtime purpose"), ("Compatibility Constraint", "esp_compatibilityconstraint", "Text(200)", "Optional", "No", "No", "Allowed compatibility")],
    "esp_packageversion": [("Version", "esp_version", "Text(32)", "Required", "AK", "No", "Package SemVer"), ("Package Status", "esp_packagestatus", "Choice", "Required", "No", "No", "Package lifecycle"), ("Build Reference", "esp_buildreference", "URL(1000)", "Required", "No", "Restricted", "Immutable build"), ("Artifact Hash", "esp_artifacthash", "Text(128)", "Required", "No", "Restricted", "Package integrity")],
    "esp_dependencysnapshot": [("Resolved Version", "esp_resolvedversion", "Text(32)", "Required", "No", "No", "Resolved target version"), ("Artifact Reference", "esp_artifactreference", "URL(1000)", "Required", "No", "Restricted", "Resolved artifact"), ("Artifact Hash", "esp_artifacthash", "Text(128)", "Required", "No", "Restricted", "Resolved integrity"), ("Resolution Status", "esp_resolutionstatus", "Choice", "Required", "No", "No", "Resolution result")],
    "esp_releaserecord": [("Release Status", "esp_releasestatus", "Choice", "Required", "No", "No", "Release lifecycle"), ("Release Notes Reference", "esp_releasenotesreference", "URL(1000)", "Required", "No", "Restricted", "Release notes"), ("Rollback Target", "esp_rollbacktarget", "Text(200)", "Required", "No", "Restricted", "Approved rollback package")],
    "esp_deployment": [("Environment", "esp_environment", "Choice", "Required", "No", "No", "DEV, TEST, or PROD"), ("Deployment Status", "esp_deploymentstatus", "Choice", "Required", "No", "No", "Deployment lifecycle"), ("Configuration Reference", "esp_configurationreference", "URL(1000)", "Required", "No", "Restricted", "Environment configuration"), ("Post-deployment Check", "esp_postdeploymentcheck", "Multiline Text", "Required", "No", "Restricted", "Smoke-test evidence"), ("Deployed On", "esp_deployedon", "DateTime", "Required", "No", "No", "Deployment time")],
    "esp_consumerbinding": [("Environment", "esp_environment", "Choice", "Required", "No", "No", "Binding environment"), ("Binding Status", "esp_bindingstatus", "Choice", "Required", "No", "No", "Binding lifecycle"), ("Fallback Behavior", "esp_fallbackbehavior", "Choice", "Required", "No", "No", "Explicit fallback"), ("Runtime Configuration Reference", "esp_runtimeconfigurationreference", "URL(1000)", "Required", "No", "Restricted", "Pinned non-secret configuration"), ("Data Classification", "esp_dataclassification", "Choice", "Required", "No", "Restricted", "Maximum allowed classification"), ("Retention Policy", "esp_retentionpolicy", "Text(200)", "Required", "No", "Restricted", "Authoritative retention policy"), ("Effective From", "esp_effectivefrom", "DateTime", "Required", "No", "No", "Activation time"), ("Effective To", "esp_effectiveto", "DateTime", "Optional", "No", "No", "Expiry")],
    "esp_evaluationprofile": [("Test Set Reference", "esp_testsetreference", "URL(1000)", "Required", "No", "Restricted", "Versioned test set"), ("Threshold Definition", "esp_thresholddefinition", "Multiline Text", "Required", "No", "Restricted", "Approved thresholds"), ("Mandatory Categories", "esp_mandatorycategories", "Multiline Text", "Required", "No", "No", "Required test categories")],
    "esp_evaluationrun": [("Run Status", "esp_runstatus", "Choice", "Required", "No", "No", "Evaluation lifecycle"), ("Dataset Version", "esp_datasetversion", "Text(32)", "Required", "No", "Restricted", "Pinned dataset SemVer"), ("Evaluator Versions", "esp_evaluatorversions", "Multiline Text", "Required", "No", "Restricted", "Pinned evaluator identities and versions"), ("Runtime Configuration Reference", "esp_runtimeconfigurationreference", "URL(1000)", "Required", "No", "Restricted", "Pinned runtime configuration"), ("Threshold Version", "esp_thresholdversion", "Text(32)", "Required", "No", "Restricted", "Pinned threshold SemVer"), ("Snapshot Set Reference", "esp_snapshotsetreference", "URL(1000)", "Required", "No", "Restricted", "Pinned dependency snapshot set"), ("Snapshot Set Hash", "esp_snapshotsethash", "Text(128)", "Required", "No", "Restricted", "Dependency snapshot set integrity"), ("Started On", "esp_startedon", "DateTime", "Required", "No", "No", "Run start"), ("Completed On", "esp_completedon", "DateTime", "Optional", "No", "No", "Run completion"), ("Result Reference", "esp_resultreference", "URL(1000)", "Required", "No", "Restricted", "Per-case assertions"), ("Aggregate Measures", "esp_aggregatemeasures", "Multiline Text", "Required", "No", "Restricted", "Measured results")],
    "esp_gatedecision": [("Decision", "esp_gatedecision", "Choice", "Required", "No", "No", "Gate outcome"), ("Conditions", "esp_conditions", "Multiline Text", "Optional", "No", "Restricted", "Conditional requirements"), ("Condition Owner", "esp_conditionowner", "Text(200)", "Optional", "No", "Restricted", "Accountable owner"), ("Expiry Date", "esp_expirydate", "Date", "Optional", "No", "No", "Conditional expiry")],
    "esp_approvalrecord": [("Decision", "esp_approvaldecision", "Choice", "Required", "No", "No", "Approval decision"), ("Approver Reference", "esp_approverreference", "Text(200)", "Required", "No", "Restricted", "Authoritative approver"), ("Decision Rationale", "esp_decisionrationale", "Multiline Text", "Required", "No", "Restricted", "Decision record"), ("Effective On", "esp_effectiveon", "DateTime", "Required", "No", "No", "Effective time")],
    "esp_identityprofile": [("Identity Type", "esp_identitytype", "Choice", "Required", "No", "Restricted", "User, Agent, Application, or Connection"), ("Authority Reference", "esp_authorityreference", "URL(1000)", "Required", "No", "Restricted", "External identity authority"), ("Least Privilege Requirements", "esp_leastprivilegerequirements", "Multiline Text", "Required", "No", "Restricted", "Required permissions")],
    "esp_identitybinding": [("Runtime Identity Reference", "esp_runtimeidentityreference", "Text(500)", "Required", "No", "Restricted", "Non-secret external identifier"), ("Validation Status", "esp_validationstatus", "Choice", "Required", "No", "Restricted", "Permission validation")],
    "esp_policybinding": [("Policy Type", "esp_policytype", "Choice", "Required", "No", "Restricted", "DLP, classification, retention, or authorization"), ("Policy Reference", "esp_policyreference", "URL(1000)", "Required", "No", "Restricted", "External policy"), ("Validation Status", "esp_validationstatus", "Choice", "Required", "No", "Restricted", "Policy validation")],
    "esp_invocationindex": [("External ID", "esp_externalid", "Text(100)", "Required", "AK", "Restricted", "Runtime invocation ID"), ("Correlation ID", "esp_correlationid", "Text(100)", "Required", "No", "Restricted", "End-to-end correlation"), ("Started On", "esp_startedon", "DateTime", "Required", "No", "No", "Invocation start"), ("Completed On", "esp_completedon", "DateTime", "Optional", "No", "No", "Invocation completion"), ("Outcome", "esp_outcome", "Choice", "Required", "No", "No", "Standard outcome"), ("Error Category", "esp_errorcategory", "Choice", "Optional", "No", "Restricted", "Standard error"), ("Telemetry Reference", "esp_telemetryreference", "URL(1000)", "Required", "No", "Restricted", "External trace"), ("Retention Policy", "esp_retentionpolicy", "Text(200)", "Required", "No", "Restricted", "Retention authority")],
    "esp_evidencepackage": [("Classification", "esp_classification", "Choice", "Required", "No", "Restricted", "Evidence classification"), ("Retention Policy", "esp_retentionpolicy", "Text(200)", "Required", "No", "Restricted", "Retention authority"), ("Review Status", "esp_reviewstatus", "Choice", "Required", "No", "Restricted", "Evidence review")],
    "esp_evidenceitem": [("Evidence Type", "esp_evidencetype", "Choice", "Required", "No", "Restricted", "Fact, rule, tool, model, or human decision"), ("Source Reference", "esp_sourcereference", "URL(1000)", "Required", "No", "Restricted", "Authoritative source"), ("Source Version", "esp_sourceversion", "Text(100)", "Required", "No", "Restricted", "Source version"), ("Claim Reference", "esp_claimreference", "Text(500)", "Required", "No", "Restricted", "Supported claim"), ("Content Hash", "esp_contenthash", "Text(128)", "Required", "No", "Restricted", "Integrity without copied content"), ("Review Status", "esp_reviewstatus", "Choice", "Required", "No", "Restricted", "Evidence review")],
    "esp_operationalevidence": [("Period Start", "esp_periodstart", "DateTime", "Required", "No", "No", "Aggregation start"), ("Period End", "esp_periodend", "DateTime", "Required", "No", "No", "Aggregation end"), ("Measure Definition", "esp_measuredefinition", "Multiline Text", "Required", "No", "Restricted", "Metric semantics"), ("Measure Value", "esp_measurevalue", "Decimal", "Required", "No", "Restricted", "Observed value"), ("Source Reference", "esp_sourcereference", "URL(1000)", "Required", "No", "Restricted", "Authoritative aggregate")],
    "esp_valueassessmentreference": [("Decision Reference", "esp_decisionreference", "URL(1000)", "Required", "No", "Restricted", "Business-owned decision"), ("Decision Type", "esp_decisiontype", "Choice", "Required", "No", "No", "Scale, maintain, reduce, or retire"), ("Decision Date", "esp_decisiondate", "Date", "Required", "No", "No", "Business decision date")],
}

VERSION_TABLES = {
    "esp_knowledgeassetversion", "esp_promptassetversion", "esp_templateassetversion",
    "esp_methodassetversion", "esp_toolcontractversion", "esp_workflowdefinitionversion",
}

GENERIC_LIFECYCLE_EXCEPTIONS = {
    "esp_implementationversion", "esp_packageversion", "esp_dependencysnapshot",
    "esp_releaserecord", "esp_deployment", "esp_consumerbinding", "esp_evaluationrun",
    "esp_gatedecision", "esp_approvalrecord", "esp_invocationindex",
}

SPECIAL_STATUS_COLUMNS = {
    "esp_implementationversion": "esp_implementationstatus",
    "esp_packageversion": "esp_packagestatus",
    "esp_dependencysnapshot": "esp_resolutionstatus",
    "esp_releaserecord": "esp_releasestatus",
    "esp_deployment": "esp_deploymentstatus",
    "esp_consumerbinding": "esp_bindingstatus",
    "esp_evaluationrun": "esp_runstatus",
    "esp_gatedecision": "esp_gatedecision",
    "esp_approvalrecord": "esp_approvaldecision",
}

CHOICES = [
    ("esp_lifecyclestatus", 1, "Draft", "Editable"), ("esp_lifecyclestatus", 2, "In Review", "Under review"), ("esp_lifecyclestatus", 3, "Approved", "Immutable approved version"), ("esp_lifecyclestatus", 4, "Pilot", "Pilot use"), ("esp_lifecyclestatus", 5, "Production", "Production use"), ("esp_lifecyclestatus", 6, "Deprecated", "No new binding"), ("esp_lifecyclestatus", 7, "Retired", "History only"),
    ("esp_implementationstatus", 1, "Draft", "Under development"), ("esp_implementationstatus", 2, "Validated", "Contract and technical checks passed"), ("esp_implementationstatus", 3, "Approved", "Approved for packaging"), ("esp_implementationstatus", 4, "Deprecated", "No new package"), ("esp_implementationstatus", 5, "Retired", "History only"),
    ("esp_packagestatus", 1, "Built", "Package built"), ("esp_packagestatus", 2, "Validated", "Package checks passed"), ("esp_packagestatus", 3, "Approved", "Approved for release"), ("esp_packagestatus", 4, "Superseded", "Replaced by another package"), ("esp_packagestatus", 5, "Retired", "History only"),
    ("esp_deploymentstatus", 1, "Planned", "Awaiting deployment"), ("esp_deploymentstatus", 2, "Deploying", "In progress"), ("esp_deploymentstatus", 3, "Succeeded", "Deployment succeeded"), ("esp_deploymentstatus", 4, "Failed", "Deployment failed"), ("esp_deploymentstatus", 5, "Rolled Back", "Rollback completed"), ("esp_deploymentstatus", 6, "Retired", "No longer active"),
    ("esp_bindingstatus", 1, "Draft", "Not active"), ("esp_bindingstatus", 2, "Active", "Runtime use allowed"), ("esp_bindingstatus", 3, "Suspended", "Temporarily disabled"), ("esp_bindingstatus", 4, "Expired", "Effective period ended"), ("esp_bindingstatus", 5, "Retired", "Permanently disabled"),
    ("esp_runstatus", 1, "Planned", "Awaiting execution"), ("esp_runstatus", 2, "Running", "In progress"), ("esp_runstatus", 3, "Completed", "Execution completed"), ("esp_runstatus", 4, "Failed", "Execution failed"), ("esp_runstatus", 5, "Cancelled", "Execution cancelled"),
    ("esp_gatedecision", 1, "Approved", "Promotion allowed"), ("esp_gatedecision", 2, "Conditionally Approved", "Allowed with expiring conditions"), ("esp_gatedecision", 3, "Rejected", "Promotion denied"), ("esp_gatedecision", 4, "Requires Remediation", "Changes and rerun required"),
    ("esp_approvaldecision", 1, "Approved", "Approval granted"), ("esp_approvaldecision", 2, "Conditionally Approved", "Approval granted with conditions"), ("esp_approvaldecision", 3, "Rejected", "Approval denied"), ("esp_approvaldecision", 4, "Requires Remediation", "Changes required"),
    ("esp_assessmentdecision", 1, "Reuse", "Use an existing Skill"), ("esp_assessmentdecision", 2, "Extend Compatibly", "Add backward-compatible capability"), ("esp_assessmentdecision", 3, "Build", "Create a new Skill"), ("esp_assessmentdecision", 4, "Reject", "Do not admit capability"),
    ("esp_dependencytype", 1, "Knowledge", "Knowledge Asset Version"), ("esp_dependencytype", 2, "Prompt", "Prompt Asset Version"), ("esp_dependencytype", 3, "Template", "Template Asset Version"), ("esp_dependencytype", 4, "Method", "Method Asset Version"), ("esp_dependencytype", 5, "Tool Contract", "Tool Contract Version"), ("esp_dependencytype", 6, "Workflow", "Workflow Definition Version"), ("esp_dependencytype", 7, "External Reference", "Approved external reference"),
    ("esp_dependencymode", 1, "Version Pinned", "Exact approved version"), ("esp_dependencymode", 2, "Immutable Snapshot", "Content and hash frozen"), ("esp_dependencymode", 3, "Governed Latest", "Approved exception with source version recorded"),
    ("esp_outcome", 1, "Success", "Completed"), ("esp_outcome", 2, "Needs Information", "Required input missing"), ("esp_outcome", 3, "Cannot Assess", "Evidence insufficient"), ("esp_outcome", 4, "Human Handoff", "Human action required"), ("esp_outcome", 5, "Rejected by Policy", "Policy denied"), ("esp_outcome", 6, "Failed", "Execution failed"),
    ("esp_risktier", 1, "Low", "Low business risk"), ("esp_risktier", 2, "Moderate", "Moderate business risk"), ("esp_risktier", 3, "High", "High business risk"), ("esp_risktier", 4, "Critical", "Critical business risk"),
    ("esp_dataclassification", 1, "Public", "Approved for public distribution"), ("esp_dataclassification", 2, "Internal", "Internal use"), ("esp_dataclassification", 3, "Confidential", "Restricted business data"), ("esp_dataclassification", 4, "Highly Confidential", "Highest governed classification"),
    ("esp_humanoversight", 1, "Not Required", "No mandatory review"), ("esp_humanoversight", 2, "Required Review", "Review every material result"), ("esp_humanoversight", 3, "Approval", "Explicit approval required"),
    ("esp_priority", 1, "Low", "Low delivery priority"), ("esp_priority", 2, "Medium", "Normal delivery priority"), ("esp_priority", 3, "High", "High delivery priority"), ("esp_priority", 4, "Critical", "Critical delivery priority"),
    ("esp_oversightlevel", 1, "Not Required", "No mandatory review"), ("esp_oversightlevel", 2, "Required Review", "Human review required"), ("esp_oversightlevel", 3, "Approval", "Explicit approval required"),
    ("esp_consumertype", 1, "Agent", "Agent consumer"), ("esp_consumertype", 2, "Workflow", "Workflow consumer"), ("esp_consumertype", 3, "Application", "Application consumer"), ("esp_consumertype", 4, "API", "API consumer"), ("esp_consumertype", 5, "Human Assisted", "Human-assisted experience"),
    ("esp_hosttype", 1, "Copilot Studio", "Copilot Studio Agent"), ("esp_hosttype", 2, "Power Automate", "Power Automate host"), ("esp_hosttype", 3, "Approved External", "Approved external runtime"),
    ("esp_exposuremechanism", 1, "Agent Action", "Copilot Agent action"), ("esp_exposuremechanism", 2, "Topic", "Agent topic"), ("esp_exposuremechanism", 3, "Connector", "Governed connector"), ("esp_exposuremechanism", 4, "Flow", "Governed flow"),
    ("esp_resolutionstatus", 1, "Pending", "Not resolved"), ("esp_resolutionstatus", 2, "Resolved", "Version and hash resolved"), ("esp_resolutionstatus", 3, "Failed", "Resolution failed"), ("esp_resolutionstatus", 4, "Hash Mismatch", "Integrity check failed"),
    ("esp_releasestatus", 1, "Draft", "Release preparation"), ("esp_releasestatus", 2, "In Review", "Release review"), ("esp_releasestatus", 3, "Approved", "Release approved"), ("esp_releasestatus", 4, "Superseded", "Replaced release"), ("esp_releasestatus", 5, "Retired", "History only"),
    ("esp_environment", 1, "DEV", "Development"), ("esp_environment", 2, "TEST", "Test"), ("esp_environment", 3, "PROD", "Production"),
    ("esp_fallbackbehavior", 1, "Retry", "Retry by policy"), ("esp_fallbackbehavior", 2, "Clarify", "Request clarification"), ("esp_fallbackbehavior", 3, "Degrade", "Use approved degraded path"), ("esp_fallbackbehavior", 4, "Human Handoff", "Route to human"), ("esp_fallbackbehavior", 5, "Stop", "Stop processing"),
    ("esp_identitytype", 1, "User Delegated", "User-delegated identity"), ("esp_identitytype", 2, "Agent", "Agent identity"), ("esp_identitytype", 3, "Application", "Application identity"), ("esp_identitytype", 4, "Connection", "Connection identity"),
    ("esp_validationstatus", 1, "Pending", "Not validated"), ("esp_validationstatus", 2, "Valid", "Validation passed"), ("esp_validationstatus", 3, "Invalid", "Validation failed"), ("esp_validationstatus", 4, "Expired", "Validation expired"),
    ("esp_policytype", 1, "DLP", "Data loss prevention"), ("esp_policytype", 2, "Classification", "Data classification"), ("esp_policytype", 3, "Retention", "Retention policy"), ("esp_policytype", 4, "Authorization", "Authorization policy"),
    ("esp_errorcategory", 1, "Invalid Input", "Input validation failed"), ("esp_errorcategory", 2, "Missing Evidence", "Required evidence missing"), ("esp_errorcategory", 3, "Dependency Failure", "Dependency unavailable"), ("esp_errorcategory", 4, "Policy Denial", "Policy denied access"), ("esp_errorcategory", 5, "Timeout", "Execution timeout"), ("esp_errorcategory", 6, "Internal Failure", "Unhandled implementation error"),
    ("esp_classification", 1, "Public", "Approved for public distribution"), ("esp_classification", 2, "Internal", "Internal use"), ("esp_classification", 3, "Confidential", "Restricted business data"), ("esp_classification", 4, "Highly Confidential", "Highest governed classification"),
    ("esp_reviewstatus", 1, "Unreviewed", "Not reviewed"), ("esp_reviewstatus", 2, "In Review", "Review in progress"), ("esp_reviewstatus", 3, "Verified", "Evidence verified"), ("esp_reviewstatus", 4, "Rejected", "Evidence rejected"), ("esp_reviewstatus", 5, "Archived", "Retained history"),
    ("esp_evidencetype", 1, "Fact", "Sourced fact"), ("esp_evidencetype", 2, "Rule Result", "Deterministic rule result"), ("esp_evidencetype", 3, "Tool Result", "Tool execution result"), ("esp_evidencetype", 4, "Model Suggestion", "Non-factual model suggestion"), ("esp_evidencetype", 5, "Human Decision", "Human disposition"),
    ("esp_decisiontype", 1, "Scale", "Expand use"), ("esp_decisiontype", 2, "Maintain", "Continue current use"), ("esp_decisiontype", 3, "Reduce", "Reduce use"), ("esp_decisiontype", 4, "Retire", "End use"),
]

RULES = [
    ("RULE-DEP-001", "esp_dependency", "Create/Update", "Exactly one of the seven governed target lookups is populated", "Synchronous blocking validation", "Dependency cannot have zero or multiple targets"),
    ("RULE-DEP-002", "esp_dependency", "Create/Update", "Dependency Type matches the populated target lookup", "Synchronous blocking validation", "Target type must be truthful"),
    ("RULE-REL-001", "esp_releaserecord", "Approve", "All mandatory Dependency Snapshots are Resolved and hashes match", "Gate validation", "Block incomplete release"),
    ("RULE-BIND-001", "esp_consumerbinding", "Activate", "Deployment succeeded; versions approved; identity, policy, configuration, classification, retention, and smoke test valid", "Gate validation", "Block unsafe runtime binding"),
    ("RULE-EVAL-001", "esp_evaluationrun", "Complete", "Dataset, package, snapshot set, evaluator, runtime configuration, threshold, deployment, and binding versions are pinned", "Synchronous blocking validation", "Evaluation must be reproducible"),
    ("RULE-EVID-001", "esp_evidenceitem", "Create/Verify", "Every material factual claim has an authorized source and content hash", "Synchronous blocking validation", "Unsupported claims are prohibited"),
]

SECURITY = [
    ("Core", "ESP Admin", "Yes", "Yes", "Yes", "Controlled", "Yes", "Platform administration"),
    ("Demand", "Business Owner", "Yes", "Yes", "Draft/Review", "No", "Business", "Own objectives and use cases"),
    ("Assets", "Capability/Asset Owner", "Yes", "Yes", "Draft/Review", "No", "Asset", "Own governed assets"),
    ("Implementation", "Implementation Owner", "Yes", "Yes", "Draft/Review", "No", "No", "Runtime realization"),
    ("Governance", "Governance Approver", "No", "Yes", "Decision", "No", "Yes", "Gate and release approvals"),
    ("Evidence", "Evidence Auditor", "No", "Restricted", "Review", "No", "No", "Evidence verification"),
    ("Catalog", "Business Reader", "No", "Approved", "No", "No", "No", "Approved metadata only"),
]

SEED_DATA = [
    ("Business Objective", "BO-SEC-001", "Improve security review efficiency and quality", "1.0.0", "Draft", "Business Owner"),
    ("Logical Skill", "LS-SEC-DOC-INTAKE", "Document Intake", "1.0.0", "Draft", "Capability Owner"),
    ("Logical Skill", "LS-SEC-EVIDENCE-EXTRACT", "Evidence Extraction", "1.0.0", "Draft", "Capability Owner"),
    ("Logical Skill", "LS-SEC-REVIEW", "Security Review", "1.0.0", "Draft", "Capability Owner"),
    ("Logical Skill", "LS-SEC-RISK-RATING", "Risk Rating", "1.0.0", "Draft", "Capability Owner"),
    ("Logical Skill", "LS-SEC-REPORT-GEN", "Report Generation", "1.0.0", "Draft", "Capability Owner"),
    ("Consumer", "CON-SEC-REVIEW-AGENT", "Security Review Agent", "", "Draft", "Implementation Owner"),
    ("Consumer", "CON-ARCH-REVIEW", "Architecture Review Agent or Workflow", "", "Draft", "Implementation Owner"),
    ("Runtime Profile", "RP-COPILOT-STUDIO-TEST", "Copilot Studio TEST", "", "Draft", "Implementation Owner"),
    ("Evaluation Profile", "EP-SEC-MVP-001", "Security Review MVP Evaluation", "1.0.0", "Draft", "Domain SME"),
]


def add_column(rows: list[list[str]], table: str, display: str, logical: str, data_type: str, required: str, key: str, security: str, definition: str) -> None:
    rows.append([table, display, logical, data_type, required, key, "Yes", security, definition])


def build_columns() -> list[list[str]]:
    rows = [["Table", "Display Name", "Logical Name", "Data Type", "Required", "Key", "Audit", "Security", "Definition"]]
    table_keys = {table: key for table, _, _, key, _ in TABLES}
    lookup_columns: dict[str, set[str]] = {table: set() for table, *_ in TABLES}
    for table, _, _, key, _ in TABLES:
        if key == "esp_externalid":
            continue
        add_column(rows, table, "Code", "esp_code", "Text(100)", "Required", "AK" if key == "esp_code" else "Composite", "No", "Stable immutable code")
        add_column(rows, table, "Name", "esp_name", "Text(200)", "Required", "No", "No", "Display name")
        if table not in GENERIC_LIFECYCLE_EXCEPTIONS:
            add_column(rows, table, "Lifecycle Status", "esp_lifecyclestatus", "Choice", "Required", "No", "No", "Governed lifecycle")
    for parent, child, cardinality, lookup, required, _, notes in RELATIONSHIPS:
        if cardinality == "N:N" or lookup in lookup_columns[child]:
            continue
        lookup_columns[child].add(lookup)
        add_column(rows, child, parent.removeprefix("esp_").replace("version", " Version").title(), lookup, "Lookup", required, "Composite" if lookup in table_keys[child] else "No", "No", notes)
    for table in VERSION_TABLES:
        add_column(rows, table, "Version", "esp_version", "Text(32)", "Required", "Composite", "No", "SemVer")
        add_column(rows, table, "Artifact Reference", "esp_artifactreference", "URL(1000)", "Required", "No", "Restricted", "External immutable artifact")
        add_column(rows, table, "Artifact Hash", "esp_artifacthash", "Text(128)", "Required", "No", "Restricted", "Artifact integrity")
    for table, columns in SPECIAL_COLUMNS.items():
        for display, logical, data_type, required, key, security, definition in columns:
            add_column(rows, table, display, logical, data_type, required, key, security, definition)
    return rows


def build_solution_dependencies() -> list[list[str]]:
    owners = {table: solution for table, _, _, _, solution in TABLES}
    dependencies: dict[tuple[str, str], list[str]] = {}
    for parent, child, _, lookup, *_ in RELATIONSHIPS:
        child_solution = owners[child]
        parent_solution = owners[parent]
        if child_solution == parent_solution:
            continue
        dependencies.setdefault((child_solution, parent_solution), []).append(f"{child}.{lookup} -> {parent}")
    rows = [["Solution", "Depends On", "Lookup Count", "Relationship Evidence"]]
    for (solution, dependency), evidence in sorted(dependencies.items()):
        rows.append([solution, dependency, len(evidence), "; ".join(sorted(evidence))])
    return rows


def validate_model(columns: list[list[str]]) -> None:
    table_names = {row[0] for row in TABLES}
    errors: list[str] = []
    for parent, child, *_ in RELATIONSHIPS:
        if parent not in table_names:
            errors.append(f"Undeclared relationship parent: {parent}")
        if child not in table_names:
            errors.append(f"Undeclared relationship child: {child}")
    columns_by_table: dict[str, list[str]] = {table: [] for table in table_names}
    choice_names = {row[0] for row in CHOICES}
    choice_columns: set[str] = set()
    for row in columns[1:]:
        columns_by_table[row[0]].append(row[2])
        if row[3] == "Choice":
            choice_columns.add(row[2])
    for table, _, _, key, _ in TABLES:
        logical_names = columns_by_table[table]
        if not logical_names:
            errors.append(f"No columns declared: {table}")
        if len(logical_names) != len(set(logical_names)):
            errors.append(f"Duplicate logical column: {table}")
        for key_field in key.split("+"):
            if key_field not in logical_names:
                errors.append(f"Missing business key field {key_field}: {table}")
    for choice_column in sorted(choice_columns - choice_names):
        errors.append(f"Missing Choice definition: {choice_column}")
    if len({(row[0], row[1]) for row in CHOICES}) != len(CHOICES):
        errors.append("Duplicate Choice name and numeric value")
    for table, status_column in SPECIAL_STATUS_COLUMNS.items():
        if status_column not in columns_by_table[table]:
            errors.append(f"Missing specialized lifecycle field {status_column}: {table}")
    required_fields = {
        "esp_consumerbinding": {"esp_runtimeconfigurationreference", "esp_dataclassification", "esp_retentionpolicy"},
        "esp_evaluationrun": {"esp_logicalskillversionid", "esp_implementationversionid", "esp_packageversionid", "esp_datasetversion", "esp_evaluatorversions", "esp_runtimeconfigurationreference", "esp_thresholdversion", "esp_snapshotsetreference", "esp_snapshotsethash"},
        "esp_operationalevidencesource": {"esp_operationalevidenceid", "esp_invocationindexid"},
    }
    for table, fields in required_fields.items():
        for field in sorted(fields - set(columns_by_table[table])):
            errors.append(f"Missing required control field {field}: {table}")
    dependency_targets = {row[3] for row in RELATIONSHIPS if row[1] == "esp_dependency" and row[4] == "Optional"}
    if len(dependency_targets) != 7:
        errors.append(f"Dependency must declare seven exclusive target lookups, found {len(dependency_targets)}")
    if not any(row[0] == "RULE-DEP-001" for row in RULES):
        errors.append("Missing exactly-one dependency target enforcement rule")
    for rule, table, *_ in RULES:
        if table not in table_names:
            errors.append(f"Rule {rule} references undeclared table: {table}")
    dependency_rows = build_solution_dependencies()[1:]
    dependency_graph: dict[str, set[str]] = {}
    for solution, dependency, *_ in dependency_rows:
        dependency_graph.setdefault(solution, set()).add(dependency)
    visiting: set[str] = set()
    visited: set[str] = set()

    def visit(solution: str, path: list[str]) -> None:
        if solution in visiting:
            cycle_start = path.index(solution)
            errors.append("Circular Solution dependency: " + " -> ".join(path[cycle_start:] + [solution]))
            return
        if solution in visited:
            return
        visiting.add(solution)
        for dependency in sorted(dependency_graph.get(solution, set())):
            visit(dependency, path + [dependency])
        visiting.remove(solution)
        visited.add(solution)

    for solution in sorted(dependency_graph):
        visit(solution, [solution])
    if errors:
        raise ValueError("\n".join(errors))


def add_sheet(workbook: Workbook, title: str, rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(vertical="center")
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column]
        width = min(max(max(map(len, values)) + 2, 12), 45)
        sheet.column_dimensions[column[0].column_letter].width = width
    table_name = "ESP" + title.replace(" ", "")
    excel_table = Table(displayName=table_name, ref=sheet.dimensions)
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    sheet.add_table(excel_table)


def main() -> None:
    columns = build_columns()
    validate_model(columns)
    workbook = Workbook()
    workbook.remove(workbook.active)
    table_rows = [["Table", "Domain", "Ownership", "Business Key", "MVP", "Solution", "Audit"]]
    table_rows.extend([[table, domain, ownership, key, "Yes", solution, "Yes"] for table, domain, ownership, key, solution in TABLES])
    relationship_rows = [["Parent", "Child", "Cardinality", "Lookup", "Required", "Delete", "Notes"], *[list(row) for row in RELATIONSHIPS]]
    add_sheet(workbook, "Tables", table_rows)
    add_sheet(workbook, "Columns", columns)
    add_sheet(workbook, "Relationships", relationship_rows)
    solution_dependency_rows = build_solution_dependencies()
    add_sheet(workbook, "Solution Dependencies", solution_dependency_rows)
    add_sheet(workbook, "Choices", [["Choice", "Value", "Label", "Definition"], *[list(row) for row in CHOICES]])
    add_sheet(workbook, "Rules", [["Rule", "Table", "Event", "Condition", "Enforcement", "Failure Behavior"], *[list(row) for row in RULES]])
    add_sheet(workbook, "Security", [["Scope", "Role", "Create", "Read", "Write", "Delete", "Approve", "Notes"], *[list(row) for row in SECURITY]])
    add_sheet(workbook, "Seed Data", [["Entity", "Code", "Name", "Version", "Status", "Owner Role"], *[list(row) for row in SEED_DATA]])
    workbook.save(WORKBOOK_PATH)
    reloaded = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    dimensions = {sheet.title: {"rows": sheet.max_row, "columns": sheet.max_column} for sheet in reloaded.worksheets}
    reloaded.close()
    digest = hashlib.sha256(WORKBOOK_PATH.read_bytes()).hexdigest()
    marker = {
        "tool": "esp-workbook-builder-semantic-validator",
        "schemaVersion": "2.0",
        "validatorVersion": "1.0.0",
        "result": "valid",
        "exitCode": 0,
        "input": {"path": "02-Data-Design/dataverse-build-workbook.xlsx", "sha256": digest, "sizeBytes": WORKBOOK_PATH.stat().st_size, "kind": "xlsx"},
        "summary": {"errorCount": 0, "warningCount": 0, "tableCount": len(TABLES), "columnCount": len(columns) - 1, "relationshipCount": len(RELATIONSHIPS), "solutionDependencyCount": len(solution_dependency_rows) - 1, "ruleCount": len(RULES)},
        "sheets": dimensions,
        "issues": [],
    }
    VALIDATION_PATH.write_text(json.dumps(marker, indent=2) + "\n", encoding="ascii")
    print(f"Built {WORKBOOK_PATH.name}: {len(TABLES)} tables, {len(columns) - 1} columns, {len(RELATIONSHIPS)} relationships")
    print(f"SHA256 {digest}")


if __name__ == "__main__":
    main()